"""
FreshMart - Config-Driven Multi-Format Data Exporter & Bank Statement Generator
=================================================================================
1. Reads settings from `config.ini` and narrations from `bank_narrations.json`.
2. Exports:
   - `store_orders.csv`
   - `razorpay_settlement_recon.csv`
   - `bank_statement_sbi.pdf` / `.xlsx` OR `bank_statement_union_bank.pdf` / `.xlsx`
3. Dynamically imputes non-Razorpay bank statement line items based on `imputed_expenses_percentage`.
4. Supports both SBI and Union Bank of India statement PDF layouts.
"""

import csv
import configparser
import json
import math
import os
import random
import sqlite3
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR = os.path.dirname(CURRENT_DIR)
DB_PATH = os.path.join(ROOT_DIR, "store.db")
OUTPUT_DIR = os.path.join(ROOT_DIR, "generated_data")
os.makedirs(OUTPUT_DIR, exist_ok=True)

CONFIG_PATH = os.path.join(CURRENT_DIR, "config.ini")
NARRATIONS_PATH = os.path.join(CURRENT_DIR, "bank_narrations.json")

ORDERS_CSV_PATH = os.path.join(OUTPUT_DIR, "store_orders.csv")
SETTLEMENT_CSV_PATH = os.path.join(OUTPUT_DIR, "razorpay_settlement_recon.csv")

SBI_PDF_PATH = os.path.join(OUTPUT_DIR, "bank_statement_sbi.pdf")
SBI_XLSX_PATH = os.path.join(OUTPUT_DIR, "bank_statement_sbi.xlsx")

UBI_PDF_PATH = os.path.join(OUTPUT_DIR, "bank_statement_union_bank.pdf")
UBI_XLSX_PATH = os.path.join(OUTPUT_DIR, "bank_statement_union_bank.xlsx")


def load_config():
    config = configparser.ConfigParser()
    if os.path.exists(CONFIG_PATH):
        config.read(CONFIG_PATH)
    return config


def load_narrations():
    if os.path.exists(NARRATIONS_PATH):
        with open(NARRATIONS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def export_store_orders_csv():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT order_id, customer_name, gross_amount, order_status, created_at
        FROM orders
        ORDER BY created_at ASC;
    """)
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        print("[WARNING] No order records found in store.db to export.")
        return []

    orders_data = [dict(r) for r in rows]

    with open(ORDERS_CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "order_id", "customer_name", "gross_amount", "order_status", "created_at"
        ])
        writer.writeheader()
        writer.writerows(orders_data)

    print(f"[SUCCESS] Exported {len(orders_data)} store orders to: {ORDERS_CSV_PATH}")
    return orders_data


def export_razorpay_settlement_csv():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute("""
        SELECT p.payment_id, p.order_id, p.amount, p.fee, p.tax, 
               COALESCE(p.tds, 0.0) AS tds, p.net_credit, 
               p.settlement_utr, p.status, o.created_at
        FROM payments p
        LEFT JOIN orders o ON p.order_id = o.order_id
        ORDER BY o.created_at ASC;
    """)
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        print("[WARNING] No payment records found in store.db to export.")
        return []

    # Build UTR date lookup for orphan refund & chargeback fallbacks
    utr_date_lookup = {}
    for r in rows:
        if r["created_at"] and r["settlement_utr"]:
            utr_date_lookup[r["settlement_utr"]] = r["created_at"]

    csv_data = []
    for idx, r in enumerate(rows, 1):
        created_dt = r["created_at"] or utr_date_lookup.get(r["settlement_utr"], "2026-09-01 10:00:00")
        settlement_id = f"setl_S{created_dt[5:7]}{created_dt[8:10]}_{idx:03d}"
        
        status_val = r["status"]
        pid = str(r["payment_id"])
        if status_val == "refunded" or pid.startswith("rfnd_"):
            txn_type = "refund"
        elif status_val == "dispute_hold" or pid.startswith("disp_"):
            txn_type = "dispute_hold"
        else:
            txn_type = "payment"

        csv_data.append({
            "settlement_id": settlement_id,
            "settlement_utr": r["settlement_utr"],
            "payment_id": r["payment_id"],
            "order_id": r["order_id"],
            "amount": f"{r['amount']:.2f}",
            "fee": f"{r['fee']:.2f}",
            "tax": f"{r['tax']:.2f}",
            "tds": f"{float(r['tds']):.2f}",
            "net_credit": f"{r['net_credit']:.2f}",
            "type": txn_type,
            "status": status_val,
            "created_at": created_dt,
            "settled_at": created_dt
        })

    with open(SETTLEMENT_CSV_PATH, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "settlement_id", "settlement_utr", "payment_id", "order_id",
            "amount", "fee", "tax", "tds", "net_credit", "type", "status",
            "created_at", "settled_at"
        ])
        writer.writeheader()
        writer.writerows(csv_data)

    print(f"[SUCCESS] Exported {len(csv_data)} gateway transactions to: {SETTLEMENT_CSV_PATH}")
    return rows


from edge_case_simulators.simulate_dropped_webhooks import apply_dropped_webhooks_simulation
from edge_case_simulators.simulate_fee_overcharges import apply_fee_overcharges_simulation
from edge_case_simulators.simulate_non_reversed_refunds import apply_non_reversed_refunds_simulation
from edge_case_simulators.simulate_chargebacks import apply_chargebacks_simulation
from edge_case_simulators.simulate_section_194o_tds import apply_section_194o_tds_simulation


def apply_edge_case_mutations(config):
    """
    Applies Post-Processing Anomaly Mutations directly on store.db using modular simulators:
    - Edge Case 1 (Dropped Webhook): simulate_dropped_webhooks.py
    - Edge Case 2 (Fee Overcharge): simulate_fee_overcharges.py
    - Edge Case 3 (Non-Reversed Refunds): simulate_non_reversed_refunds.py
    - Edge Case 4 (Chargeback Holds): simulate_chargebacks.py
    - Edge Case 5 (Section 194-O TDS): simulate_section_194o_tds.py
    """
    enable_edge_cases = config.getboolean("EDGE_CASES", "enable_edge_cases", fallback=True)
    if not enable_edge_cases:
        print("[MUTATOR] Edge Cases disabled in config.ini. Exporting 100% matched clean baseline data.")
        return

    dropped_count = config.getint("EDGE_CASES", "dropped_webhook_count", fallback=2)
    fee_overcharge_count = config.getint("EDGE_CASES", "fee_overcharge_count", fallback=2)
    orphan_count = config.getint("EDGE_CASES", "orphan_refund_count", fallback=2)
    chargeback_count = config.getint("EDGE_CASES", "chargeback_hold_count", fallback=1)

    base_mdr = config.getfloat("CONTRACTED_RATES", "mdr_rate_percent", fallback=2.0) / 100.0
    gst_rate = config.getfloat("CONTRACTED_RATES", "gst_rate_percent", fallback=18.0) / 100.0

    is_tds_str = config.get("MERCHANT_TAX_PROFILE", "is_tds_applicable", fallback="no").strip().lower()
    is_tds_applicable = is_tds_str in ["yes", "true", "1", "y"]
    tds_rate = config.getfloat("MERCHANT_TAX_PROFILE", "tds_rate_percent", fallback=1.0) / 100.0
    gstin = config.get("MERCHANT_TAX_PROFILE", "gstin", fallback="36AATUF1234F1ZV")
    pan = config.get("MERCHANT_TAX_PROFILE", "pan", fallback="ABCDE1234F")

    print("\n[MUTATOR] Applying Config-Driven Isolated Edge Case Simulations...")

    # 1. Edge Case 1: Dropped Webhooks
    apply_dropped_webhooks_simulation(DB_PATH, dropped_count)

    # 2. Edge Case 2: Gateway Fee Overcharges
    apply_fee_overcharges_simulation(DB_PATH, fee_overcharge_count, base_mdr=base_mdr, gst_rate=gst_rate)

    # 3. Edge Case 3: Dynamic Non-Reversed Customer Refunds
    apply_non_reversed_refunds_simulation(DB_PATH, orphan_count, base_mdr=base_mdr, gst_rate=gst_rate)

    # 4. Edge Case 4: Customer Chargebacks & Dispute Fee Holds
    apply_chargebacks_simulation(DB_PATH, chargeback_count, dispute_fee=500.0, gst_rate=gst_rate)

    # 5. Edge Case 5: Section 194-O Statutory TDS Upfront Deduction
    apply_section_194o_tds_simulation(DB_PATH, is_tds_applicable=is_tds_applicable, tds_rate=tds_rate, gstin=gstin, pan=pan)


def prepare_bank_transactions(payment_rows, config):
    imputed_pct = config.getint("SIMULATION", "imputed_expenses_percentage", fallback=25)
    start_date_str = config.get("SIMULATION", "start_date", fallback="2026-09-01")
    end_date_str = config.get("SIMULATION", "end_date", fallback="2026-09-23")

    narrations_pool = load_narrations()
    razorpay_count = len(payment_rows)
    impute_count = math.ceil(razorpay_count * (imputed_pct / 100.0))

    statement_txns = []

    # 1. Group Razorpay Payouts by Settlement UTR Batch
    from collections import defaultdict
    utr_batches = defaultdict(lambda: {"net_credit": 0.0, "txn_date": "", "utr": ""})

    for r in payment_rows:
        dt_str = (r["created_at"] or start_date_str)[:10]
        utr = r["settlement_utr"]
        net_credit = float(r["net_credit"])
        utr_batches[utr]["net_credit"] += net_credit
        utr_batches[utr]["txn_date"] = dt_str
        utr_batches[utr]["utr"] = utr

    for utr, batch in utr_batches.items():
        narration = f"CMS/RAZORPAY/BATCH_SETL/{utr}/CR"
        statement_txns.append({
            "txn_date": batch["txn_date"],
            "value_date": batch["txn_date"],
            "description": narration,
            "ref_no": "-",
            "branch_code": "04051",
            "debit": 0.00,
            "credit": round(batch["net_credit"], 2),
            "type": "CR"
        })

    # 2. Impute non-Razorpay Bank Expenses dynamically from bank_narrations.json
    if narrations_pool:
        if impute_count <= len(narrations_pool):
            selected_items = random.sample(narrations_pool, k=impute_count)
        else:
            selected_items = random.choices(narrations_pool, k=impute_count)

        try:
            start_dt = datetime.strptime(start_date_str, "%Y-%m-%d")
            end_dt = datetime.strptime(end_date_str, "%Y-%m-%d")
        except Exception:
            start_dt = datetime(2026, 9, 1)
            end_dt = datetime(2026, 9, 23)

        day_span = max((end_dt - start_dt).days, 1)

        for item in selected_items:
            random_days = random.randint(0, day_span)
            txn_dt = start_dt + timedelta(days=random_days)
            dt_str = txn_dt.strftime("%Y-%m-%d")

            amt = round(random.uniform(item["min_amount"], item["max_amount"]), 2)
            is_debit = (item["type"] == "DR")

            statement_txns.append({
                "txn_date": dt_str,
                "value_date": dt_str,
                "description": item["narration"],
                "ref_no": f"NEFT{random.randint(10000000, 99999999)}",
                "branch_code": "04051",
                "debit": amt if is_debit else 0.00,
                "credit": 0.00 if is_debit else amt,
                "type": "DR" if is_debit else "CR"
            })

    # Sort chronologically
    statement_txns.sort(key=lambda x: x["txn_date"])

    # Compute running balances from config.ini opening_balance
    opening_balance = config.getfloat("SIMULATION", "opening_balance", fallback=10000.00)
    running_balance = opening_balance
    final_txns = []

    for t in statement_txns:
        if t["type"] == "CR":
            running_balance += t["credit"]
        else:
            running_balance -= t["debit"]

        try:
            dt_obj = datetime.strptime(t["txn_date"], "%Y-%m-%d")
            formatted_date = dt_obj.strftime("%d-%m-%Y")
        except Exception:
            formatted_date = t["txn_date"]

        final_txns.append({
            "txn_date": formatted_date,
            "value_date": formatted_date,
            "description": t["description"],
            "ref_no": t["ref_no"],
            "branch_code": t["branch_code"],
            "debit": t["debit"],
            "credit": t["credit"],
            "balance": round(running_balance, 2)
        })

    return final_txns, impute_count, opening_balance


def generate_union_bank_statement_pdf(final_txns, opening_balance):
    """Generates official Union Bank of India Statement PDF matching attached screenshot structure."""
    doc = SimpleDocTemplate(
        UBI_PDF_PATH,
        pagesize=letter,
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('UBITitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=14, textColor=colors.HexColor('#cc0000'), spaceAfter=2)
    sub_style = ParagraphStyle('UBISub', parent=styles['Normal'], fontName='Helvetica', fontSize=8, textColor=colors.HexColor('#003366'), leading=10)
    cell_style = ParagraphStyle('UBICell', parent=styles['Normal'], fontName='Helvetica', fontSize=6.5, leading=8.5)

    elements = []

    # Union Bank Header
    elements.append(Paragraph("<b>यूनियन बैंक ऑफ इंडिया &bull; Union Bank of India</b>", title_style))
    elements.append(Paragraph("A Government of India Undertaking &bull; Koramangala Main Branch, Bengaluru", sub_style))
    elements.append(Paragraph("<b>Account Name:</b> FreshMart Retail Pvt Ltd &nbsp;|&nbsp; <b>Account No:</b> 0730XXXXXXXX0777 &nbsp;|&nbsp; <b>IFSC:</b> UBIN0540510", sub_style))
    elements.append(Spacer(1, 8))

    # Table Header: 7 Columns matching screenshot
    table_data = [[
        "SI", "Date", "Particulars", "Chq Num", "Withdrawal", "Deposit", "Balance"
    ]]

    total_debits = sum(t["debit"] for t in final_txns)
    total_credits = sum(t["credit"] for t in final_txns)
    closing_balance = final_txns[-1]["balance"] if final_txns else opening_balance

    for idx, t in enumerate(final_txns, start=1):
        debit_str = f"{t['debit']:,.2f}" if t['debit'] > 0 else ""
        credit_str = f"{t['credit']:,.2f}" if t['credit'] > 0 else ""
        balance_str = f"{t['balance']:,.2f} Cr" if t['balance'] >= 0 else f"{abs(t['balance']):,.2f} Dr"

        table_data.append([
            str(idx),
            t["txn_date"],
            Paragraph(t["description"], cell_style),
            "",
            debit_str,
            credit_str,
            balance_str
        ])

    col_widths = [30, 65, 230, 50, 65, 65, 75]
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#cce6ff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#002b66')),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7.5),
        ('ALIGN', (0, 0), (1, -1), 'CENTER'),
        ('ALIGN', (3, 0), (3, -1), 'CENTER'),
        ('ALIGN', (4, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#b3d1ff')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f7faff')]),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 6.5),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    elements.append(t)
    elements.append(Spacer(1, 10))

    # Summary Table matching screenshot
    summary_data = [
        ["Summary :", f"Total Debits : {total_debits:,.2f}", f"Opening Balance : {opening_balance:,.2f} Cr"],
        ["", f"Total Credits : {total_credits:,.2f}", f"Closing Balance : {closing_balance:,.2f} Cr"]
    ]
    sum_t = Table(summary_data, colWidths=[120, 230, 230])
    sum_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#cce6ff')),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7.5),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#002b66')),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#99c2ff')),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(sum_t)

    doc.build(elements)
    print(f"[SUCCESS] Generated Union Bank of India PDF Statement to: {UBI_PDF_PATH}")


def generate_union_bank_statement_xlsx(final_txns, opening_balance):
    """Generates official Union Bank of India Statement Excel Sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Union_Bank_Statement"

    headers = ["SI", "Date", "Particulars", "Chq Num", "Withdrawal", "Deposit", "Balance"]
    ws.append(headers)

    header_fill = PatternFill(start_color="CCE6FF", end_color="CCE6FF", fill_type="solid")
    header_font = Font(name="Segoe UI", size=10, bold=True, color="002B66")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin_border = Border(
        left=Side(style='thin', color='B3D1FF'),
        right=Side(style='thin', color='B3D1FF'),
        top=Side(style='thin', color='B3D1FF'),
        bottom=Side(style='thin', color='B3D1FF')
    )

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for idx, t in enumerate(final_txns, start=1):
        debit_val = t['debit'] if t['debit'] > 0 else None
        credit_val = t['credit'] if t['credit'] > 0 else None
        balance_str = f"{t['balance']:,.2f} Cr"

        ws.append([idx, t["txn_date"], t["description"], "", debit_val, credit_val, balance_str])
        r_idx = idx + 1
        for col_idx in range(1, 8):
            cell = ws.cell(row=r_idx, column=col_idx)
            cell.font = Font(name="Segoe UI", size=9)
            cell.border = thin_border
            if col_idx in [5, 6]:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", vertical="center")

    wb.save(UBI_XLSX_PATH)
    print(f"[SUCCESS] Generated Union Bank of India Excel Statement to: {UBI_XLSX_PATH}")


def generate_sbi_bank_statement_pdf(final_txns, opening_balance):
    """Generates official SBI Bank Statement PDF."""
    doc = SimpleDocTemplate(SBI_PDF_PATH, pagesize=landscape(letter), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('BankTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=13, textColor=colors.HexColor('#003366'), spaceAfter=2)
    sub_style = ParagraphStyle('BankSub', parent=styles['Normal'], fontName='Helvetica', fontSize=7.5, textColor=colors.HexColor('#333333'), leading=10)
    cell_style = ParagraphStyle('CellText', parent=styles['Normal'], fontName='Helvetica', fontSize=6.5, leading=8.5)

    elements = []
    elements.append(Paragraph("STATE BANK OF INDIA", title_style))
    elements.append(Paragraph("Commercial Branch &bull; 80 Feet Road, Koramangala 4th Block, Bengaluru - 560034", sub_style))
    elements.append(Paragraph("<b>Account Name:</b> FreshMart Retail Private Limited &nbsp;|&nbsp; <b>Account No:</b> 30981249821", sub_style))
    elements.append(Spacer(1, 8))

    table_data = [["Txn Date", "Value Date", "Description", "Ref No./Cheque No.", "Branch Code", "Debit", "Credit", "Balance"]]
    for t in final_txns:
        debit_str = f"{t['debit']:,.2f}" if t['debit'] > 0 else "-"
        credit_str = f"{t['credit']:,.2f}" if t['credit'] > 0 else "-"
        table_data.append([t["txn_date"], t["value_date"], Paragraph(t["description"], cell_style), t["ref_no"], t["branch_code"], debit_str, credit_str, f"{t['balance']:,.2f}"])

    col_widths = [62, 62, 250, 95, 55, 70, 70, 80]
    t = Table(table_data, colWidths=col_widths)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#003366')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
    ]))
    elements.append(t)
    doc.build(elements)
    print(f"[SUCCESS] Generated Official SBI Bank Statement PDF to: {SBI_PDF_PATH}")


def generate_sbi_bank_statement_xlsx(final_txns, opening_balance):
    """Generates official SBI Bank Statement Excel sheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SBI_Bank_Statement"
    headers = ["Txn Date", "Value Date", "Description", "Ref No./Cheque No.", "Branch Code", "Debit", "Credit", "Balance"]
    ws.append(headers)

    yellow_fill = PatternFill(start_color="FFF066", end_color="FFF066", fill_type="solid")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = yellow_fill

    for t in final_txns:
        ws.append([t["txn_date"], t["value_date"], t["description"], t["ref_no"], t["branch_code"], t["debit"] or None, t["credit"] or None, t["balance"]])

    wb.save(SBI_XLSX_PATH)
    print(f"[SUCCESS] Generated SBI Bank Statement Excel Sheet to: {SBI_XLSX_PATH}")


def main():
    config = load_config()
    bank_format = config.get("SIMULATION", "bank_pdf_format", fallback="UNION_BANK").upper()

    print("=================================================================")
    print(" [EXPORT] FreshMart Config-Driven Multi-Format Data Exporter")
    print(f" Output Directory: {OUTPUT_DIR}")
    print(f" Configured Bank PDF Format: {bank_format}")
    print("=================================================================\n")

    # Apply Edge Case Anomaly Mutations on store.db if enabled in config.ini
    apply_edge_case_mutations(config)

    export_store_orders_csv()
    payment_rows = export_razorpay_settlement_csv()

    if payment_rows:
        final_txns, impute_count, opening_bal = prepare_bank_transactions(payment_rows, config)
        if bank_format == "UNION_BANK":
            generate_union_bank_statement_pdf(final_txns, opening_bal)
            generate_union_bank_statement_xlsx(final_txns, opening_bal)
        else:
            generate_sbi_bank_statement_pdf(final_txns, opening_bal)
            generate_sbi_bank_statement_xlsx(final_txns, opening_bal)

    print("\n[COMPLETE] All reconciliation datasets successfully generated and synchronized!")


if __name__ == "__main__":
    main()
